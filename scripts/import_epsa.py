"""
import_epsa.py — Importa productos del Excel EPSA a PostgreSQL y descarga imágenes.

Mapeo de columnas Excel → BD real:
  SKU          → sku (varchar, PK)
  DESCRIPCIÓN  → nombre (varchar)
  MARCA        → incluida en descripcion
  PRECIO c/IVA → precio (numeric)
  PRECIO NETO  → precio_compra (numeric)
  DISPONIBILIDAD → activo (bool) + stock_actual (int)
  Hoja (cat.)  → categoria_id (FK → categorias.id)

Uso:
    python scripts/import_epsa.py [--solo-import] [--solo-imagenes]
"""

import sys
import os
import re
import time
import logging
import argparse
import shutil
import tempfile
import requests
import psycopg2
import openpyxl
from pathlib import Path
from io import BytesIO
from PIL import Image
from icrawler.builtin import BingImageCrawler

# ─── Configuración ────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = BASE_DIR / "data" / "FERROMAX_Catalogo_EPSA_Clasificado.xlsx"
IMG_DIR = BASE_DIR / "img" / "productos"
IMG_DIR.mkdir(parents=True, exist_ok=True)

DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "dbname": "ferromax_db",
    "user": "postgres",
    "password": "doky2010",
}

PROVEEDOR_NOMBRE = "Extra Power S.A."
HOJA_INDICE = "📋 Índice"
IMG_SIZE = (800, 800)      # 800×800: cubre retina 2× para tarjetas ~224px y zoom
IMG_BG_COLOR = (255, 255, 255)  # fondo blanco — compatible con object-contain
DELAY_BUSQUEDA = 1.5
MAX_INTENTOS_IMG = 3
MAX_PRODUCTOS_POR_CATEGORIA = 20

DISPONIBILIDAD_A_STOCK = {
    "STOCK ALTO": (True, 50),
    "STOCK MEDIO": (True, 10),
    "STOCK BAJO": (True, 3),
    "ÚLTIMA UNIDAD": (True, 1),
    "SIN STOCK": (False, 0),
    "CONSULTAR": (True, 5),
}

# ─── Logging ──────────────────────────────────────────────────────────────────
_stream_handler = logging.StreamHandler(
    stream=open(sys.stdout.fileno(), mode="w", encoding="utf-8", errors="replace", closefd=False)
)
_stream_handler.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-7s  %(message)s", datefmt="%H:%M:%S"))
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        _stream_handler,
        logging.FileHandler(BASE_DIR / "scripts" / "import_epsa.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


# ─── Helpers de BD ────────────────────────────────────────────────────────────

def get_connection():
    return psycopg2.connect(**DB_CONFIG)


def obtener_o_crear_proveedor(cur) -> int:
    cur.execute("SELECT id FROM proveedores WHERE nombre = %s", (PROVEEDOR_NOMBRE,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute(
        "INSERT INTO proveedores (nombre, created_at, updated_at) VALUES (%s, NOW(), NOW()) RETURNING id",
        (PROVEEDOR_NOMBRE,),
    )
    pid = cur.fetchone()[0]
    log.info("Proveedor '%s' creado con id=%d", PROVEEDOR_NOMBRE, pid)
    return pid


def obtener_o_crear_categoria(cur, nombre: str, cache: dict) -> int:
    if nombre in cache:
        return cache[nombre]
    cur.execute("SELECT id FROM categorias WHERE nombre = %s", (nombre,))
    row = cur.fetchone()
    if row:
        cache[nombre] = row[0]
        return row[0]
    cur.execute(
        "INSERT INTO categorias (nombre, created_at, updated_at) VALUES (%s, NOW(), NOW()) RETURNING id",
        (nombre,),
    )
    cid = cur.fetchone()[0]
    cache[nombre] = cid
    log.info("Categoría '%s' creada con id=%d", nombre, cid)
    return cid


def disponibilidad_a_campos(disp: str) -> tuple[bool, int]:
    disp_upper = (disp or "").strip().upper()
    for key, val in DISPONIBILIDAD_A_STOCK.items():
        if key in disp_upper:
            return val
    # Si no coincide, interpretar como stock disponible
    return (True, 5)


def upsert_producto(cur, p: dict) -> str:
    cur.execute("SELECT sku FROM productos WHERE sku = %s", (p["sku"],))
    existe = cur.fetchone() is not None

    cur.execute(
        """
        INSERT INTO productos (
            sku, nombre, descripcion, precio, precio_compra,
            activo, stock_actual, stock_minimo,
            categoria_id, proveedor_id, imagen_url,
            created_at, updated_at
        ) VALUES (
            %(sku)s, %(nombre)s, %(descripcion)s, %(precio)s, %(precio_compra)s,
            %(activo)s, %(stock_actual)s, %(stock_minimo)s,
            %(categoria_id)s, %(proveedor_id)s, %(imagen_url)s,
            NOW(), NOW()
        )
        ON CONFLICT (sku) DO UPDATE SET
            nombre        = EXCLUDED.nombre,
            descripcion   = EXCLUDED.descripcion,
            precio        = EXCLUDED.precio,
            precio_compra = EXCLUDED.precio_compra,
            activo        = EXCLUDED.activo,
            stock_actual  = EXCLUDED.stock_actual,
            categoria_id  = EXCLUDED.categoria_id,
            proveedor_id  = EXCLUDED.proveedor_id,
            updated_at    = NOW()
        """,
        p,
    )
    return "update" if existe else "insert"


# ─── Helpers de Excel ─────────────────────────────────────────────────────────

def es_fila_subcategoria(row) -> str | None:
    val = row[0]
    if val and isinstance(val, str):
        texto = val.strip()
        if "▾" in texto or "▸" in texto or "▶" in texto:
            nombre = re.sub(r"[▸▾▶►◆●▲▼✦\s]+", " ", texto).strip()
            nombre = re.sub(r"\s*\(\d+\s*art[íi]culos?\)", "", nombre, flags=re.IGNORECASE).strip()
            return nombre if nombre else None
    return None


def es_fila_encabezado(row) -> bool:
    val = str(row[0]).strip().upper() if row[0] else ""
    return val in ("SKU", "#", "CÓDIGO", "CODIGO")


def limpiar_precio(valor) -> float | None:
    if valor is None:
        return None
    try:
        return round(float(valor), 2)
    except (ValueError, TypeError):
        return None


def parsear_hoja(ws, nombre_categoria: str, categoria_id: int, proveedor_id: int) -> list[dict]:
    productos = []
    subcategoria_actual = ""

    for row in ws.iter_rows(min_row=1, values_only=True):
        subcat = es_fila_subcategoria(row)
        if subcat:
            subcategoria_actual = subcat
            continue

        if not row[0] or es_fila_encabezado(row):
            continue

        sku = str(row[0]).strip() if row[0] else None
        if not sku or len(sku) < 2:
            continue

        nombre = str(row[1]).strip() if row[1] else ""
        marca = str(row[2]).strip() if row[2] else ""
        precio = limpiar_precio(row[3])
        precio_compra = limpiar_precio(row[4])
        disp_texto = str(row[5]).strip() if row[5] else "SIN STOCK"

        if not nombre or precio is None:
            continue

        activo, stock_actual = disponibilidad_a_campos(disp_texto)

        # Construir descripción con marca y subcategoría
        desc_parts = []
        if marca:
            desc_parts.append(f"Marca: {marca}")
        if subcategoria_actual:
            desc_parts.append(f"Subcategoría: {subcategoria_actual}")
        desc_parts.append(f"Disponibilidad: {disp_texto}")
        descripcion = " | ".join(desc_parts)

        productos.append({
            "sku": sku,
            "nombre": nombre,
            "descripcion": descripcion,
            "precio": precio,
            "precio_compra": precio_compra,
            "activo": activo,
            "stock_actual": stock_actual,
            "stock_minimo": 0,
            "categoria_id": categoria_id,
            "proveedor_id": proveedor_id,
            "imagen_url": None,
            # para uso en búsqueda de imágenes
            "_marca": marca,
        })

    return productos


# ─── PARTE A: Importación ─────────────────────────────────────────────────────

def parte_a_importar() -> int:
    log.info("=" * 60)
    log.info("PARTE A — Importación de productos desde Excel")
    log.info("=" * 60)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    hojas_datos = [s for s in wb.sheetnames if s != HOJA_INDICE]
    log.info("Hojas encontradas: %d", len(hojas_datos))

    con = get_connection()
    cur = con.cursor()

    proveedor_id = obtener_o_crear_proveedor(cur)
    con.commit()

    cat_cache: dict[str, int] = {}
    total_insert = 0
    total_update = 0
    total_error = 0

    for nombre_hoja in hojas_datos:
        ws = wb[nombre_hoja]
        categoria_id = obtener_o_crear_categoria(cur, nombre_hoja, cat_cache)
        con.commit()

        productos = parsear_hoja(ws, nombre_hoja, categoria_id, proveedor_id)
        log.info("Hoja '%-32s' → %d productos", nombre_hoja, len(productos))

        for p in productos:
            p_db = {k: v for k, v in p.items() if not k.startswith("_")}
            try:
                accion = upsert_producto(cur, p_db)
                if accion == "insert":
                    total_insert += 1
                else:
                    total_update += 1
            except Exception as e:
                log.warning("  Error SKU=%s: %s", p.get("sku"), e)
                con.rollback()
                total_error += 1
                continue

        con.commit()

    cur.close()
    con.close()
    wb.close()

    total = total_insert + total_update
    log.info("")
    log.info("─── Resumen importación ───────────────────────────────")
    log.info("  Insertados  : %d", total_insert)
    log.info("  Actualizados: %d", total_update)
    log.info("  Errores     : %d", total_error)
    log.info("  TOTAL en BD : %d", total)
    return total


# ─── PARTE B: Imágenes ────────────────────────────────────────────────────────

def construir_query(marca: str, nombre: str) -> str:
    if marca.upper() in ("NACIONAL", "S/MARCA", "SIN MARCA", ""):
        return f"{nombre} herramienta"
    return f"{marca} {nombre} ferretería"


def letterbox_y_guardar(img: Image.Image, ruta: Path):
    """
    Escala la imagen para que quepa completa en IMG_SIZE con fondo blanco.
    No recorta: el producto siempre se ve completo, centrado.
    Resultado: JPEG 800×800 blanco, compatible con object-contain del frontend.
    """
    img_rgba = img.convert("RGBA")
    target_w, target_h = IMG_SIZE
    src_w, src_h = img_rgba.size
    escala = min(target_w / src_w, target_h / src_h)
    nuevo_w = max(1, int(src_w * escala))
    nuevo_h = max(1, int(src_h * escala))

    img_scaled = img_rgba.resize((nuevo_w, nuevo_h), Image.LANCZOS)
    canvas = Image.new("RGB", IMG_SIZE, IMG_BG_COLOR)
    offset_x = (target_w - nuevo_w) // 2
    offset_y = (target_h - nuevo_h) // 2
    canvas.paste(img_scaled, (offset_x, offset_y), mask=img_scaled.split()[3])
    canvas.save(ruta, "JPEG", quality=87, optimize=True)


def buscar_y_guardar_imagen(sku: str, marca: str, nombre: str) -> bool:
    ruta = IMG_DIR / f"{sku}.jpg"
    if ruta.exists():
        log.debug("SKU=%s ya tiene imagen, se omite.", sku)
        return True

    query = construir_query(marca, nombre)
    sku_safe = re.sub(r'[\\/:*?"<>|]', '_', sku)
    tmp_dir = Path(tempfile.mkdtemp(prefix=f"ferromax_{sku_safe}_"))

    try:
        for intento in range(1, MAX_INTENTOS_IMG + 1):
            try:
                # icrawler suprime su propio logging a nivel WARNING para no saturar consola
                crawler = BingImageCrawler(
                    storage={"root_dir": str(tmp_dir)},
                    log_level=logging.ERROR,
                )
                crawler.crawl(keyword=query, max_num=3, file_idx_offset=0)

                archivos = sorted(tmp_dir.glob("*.jpg")) + sorted(tmp_dir.glob("*.png")) + \
                           sorted(tmp_dir.glob("*.jpeg")) + sorted(tmp_dir.glob("*.webp"))

                for archivo in archivos:
                    try:
                        img = Image.open(archivo)
                        letterbox_y_guardar(img, ruta)
                        log.info("  SKU=%-12s OK  %s", sku, ruta.name)
                        return True
                    except Exception:
                        continue

                if intento < MAX_INTENTOS_IMG:
                    log.warning("  SKU=%-12s intento %d/%d sin imagen válida, reintentando...",
                                sku, intento, MAX_INTENTOS_IMG)
                    time.sleep(DELAY_BUSQUEDA * intento)
                    # limpiar temporales para el siguiente intento
                    for f in tmp_dir.iterdir():
                        f.unlink()

            except Exception as e:
                log.warning("  SKU=%-12s intento %d/%d error: %s", sku, intento, MAX_INTENTOS_IMG, e)
                if intento < MAX_INTENTOS_IMG:
                    time.sleep(DELAY_BUSQUEDA * intento)

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    log.warning("  SKU=%-12s FALLIDA después de %d intentos  query='%s'",
                sku, MAX_INTENTOS_IMG, query)
    return False


def parte_b_imagenes() -> tuple[int, int, list[str]]:
    log.info("")
    log.info("=" * 60)
    log.info("PARTE B — Descarga de imágenes")
    log.info("=" * 60)

    con = get_connection()
    cur = con.cursor()

    cur.execute(
        """
        SELECT DISTINCT c.nombre
        FROM categorias c
        JOIN productos p ON p.categoria_id = c.id
        ORDER BY c.nombre
        """
    )
    categorias = [r[0] for r in cur.fetchall()]
    log.info("Categorías en BD: %d", len(categorias))

    descargadas = 0
    fallidas = 0
    skus_fallidos: list[str] = []

    for cat in categorias:
        cur.execute(
            """
            SELECT p.sku, p.nombre, p.descripcion
            FROM productos p
            JOIN categorias c ON c.id = p.categoria_id
            WHERE c.nombre = %s
            ORDER BY p.nombre ASC
            LIMIT %s
            """,
            (cat, MAX_PRODUCTOS_POR_CATEGORIA),
        )
        filas = cur.fetchall()
        log.info("Categoría '%-32s' → %d productos", cat, len(filas))

        for sku, nombre, descripcion in filas:
            # Extraer marca de la descripcion (formato "Marca: X | ...")
            marca = ""
            if descripcion:
                m = re.search(r"Marca:\s*([^|]+)", descripcion)
                if m:
                    marca = m.group(1).strip()

            ya_existe = (IMG_DIR / f"{sku}.jpg").exists()
            ok = buscar_y_guardar_imagen(sku, marca, nombre)

            if ok:
                if not ya_existe:
                    ruta_relativa = f"./img/productos/{sku}.jpg"
                    try:
                        cur.execute(
                            "UPDATE productos SET imagen_url = %s, updated_at = NOW() WHERE sku = %s",
                            (ruta_relativa, sku),
                        )
                        con.commit()
                    except Exception as e:
                        log.warning("  No se pudo actualizar imagen_url SKU=%s: %s", sku, e)
                        con.rollback()
                descargadas += 1
            else:
                fallidas += 1
                skus_fallidos.append(sku)

            if not ya_existe:
                time.sleep(DELAY_BUSQUEDA)

    cur.close()
    con.close()
    return descargadas, fallidas, skus_fallidos


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Importa productos EPSA y descarga imágenes.")
    parser.add_argument("--solo-import", action="store_true", help="Solo importar a BD")
    parser.add_argument("--solo-imagenes", action="store_true", help="Solo descargar imágenes")
    args = parser.parse_args()

    t_inicio = time.time()
    total_productos = 0
    descargadas = fallidas = 0
    skus_fallidos: list[str] = []

    if not args.solo_imagenes:
        total_productos = parte_a_importar()

    if not args.solo_import:
        descargadas, fallidas, skus_fallidos = parte_b_imagenes()

    t_total = time.time() - t_inicio

    print()
    print("=" * 60)
    print("  RESUMEN FINAL")
    print("=" * 60)
    if not args.solo_imagenes:
        print(f"  Total productos importados : {total_productos}")
    if not args.solo_import:
        print(f"  Imágenes descargadas       : {descargadas}")
        print(f"  Imágenes fallidas          : {fallidas}")
        if skus_fallidos:
            print(f"  SKUs fallidos              : {', '.join(skus_fallidos)}")
    print(f"  Tiempo total               : {t_total:.1f}s")
    print("=" * 60)


if __name__ == "__main__":
    main()
